"""
Daily Job Search Agent - ULTRA COMPACT VERSION
- Minimal spacing, smaller fonts, condensed layout
- All companies from Excel + job search
- Platform aggregators under each role
- Reads Excel tracker DAILY
"""

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from html import unescape
import openpyxl
import re
import shutil
import tempfile
import os
import json
import requests
import time

# ============= CONFIGURATION =============
# Import configuration from config.py (create from config.template.py)
try:
    from config import EMAIL_CONFIG, TRACKER_FILE
except ImportError:
    print("ERROR: config.py not found!")
    print("Please copy config.template.py to config.py and fill in your details.")
    print("See README.md for instructions.")
    exit(1)

try:
    from config import HOT_JOB_QUERIES
except ImportError:
    HOT_JOB_QUERIES = None

# ============= LOCATION SETTINGS =============
# CUSTOMIZE THESE FOR YOUR LOCATION!
# These variables are used in job search URLs and email headers
LOCATION_CITY = "Paris"
LOCATION_COUNTRY = "France"
# Examples: "London", "UK" | "New York", "USA" | "Berlin", "Germany"

# NOTE: Company data is now dynamically built from your Excel tracker
# See build_companies_by_role() function below

# Platform aggregators by role
# Uses LOCATION_CITY and LOCATION_COUNTRY variables defined above
PLATFORM_AGGREGATORS = {
    'Software Developer (Java) - SENIOR/EXPERT': [
        {
            'name': 'Glassdoor',
            'links': [
                (f'67 Senior Java {LOCATION_CITY}', 'https://www.glassdoor.com/Job/paris-senior-java-developer-jobs-SRCH_IL.0,5_IC2881970_KO6,27.htm'),
                (f'35 Senior Java {LOCATION_COUNTRY}', 'https://www.glassdoor.com/Job/france-senior-software-engineer-java-developer-jobs-SRCH_IL.0,6_IN86_KO7,46.htm')
            ]
        },
        {
            'name': 'LinkedIn',
            'links': [
                (f'2,000+ Senior Java {LOCATION_CITY}', 'https://www.linkedin.com/jobs/java-software-engineer-jobs-paris')
            ]
        },
        {
            'name': 'EnglishJobs.fr',
            'links': [
                (f'Senior Java {LOCATION_COUNTRY}', 'https://englishjobs.fr/in/paris/java')
            ]
        },
        {
            'name': 'WelcomeToTheJungle',
            'links': [
                (f'Senior Java {LOCATION_CITY} (FR)', 'https://www.welcometothejungle.com/fr/jobs?query=java&aroundQuery=Paris')
            ]
        }
    ],
    'Backend Java Developer - SENIOR': [
        {
            'name': 'Glassdoor',
            'links': [
                (f'67 Lead Java {LOCATION_CITY}', 'https://www.glassdoor.com/Job/paris-lead-java-developer-jobs-SRCH_IL.0,5_IC2881970_KO6,25.htm')
            ]
        },
        {
            'name': 'EnglishJobs.fr',
            'links': [
                (f'Backend Developer {LOCATION_COUNTRY}', 'https://englishjobs.fr/jobs/backend_developer')
            ]
        },
         {
            'name': 'WelcomeToTheJungle',
            'links': [
                (f'Backend Java {LOCATION_CITY} (FR)', 'https://www.welcometothejungle.com/fr/jobs?query=backend+java&aroundQuery=Paris')
            ]
        }
    ],
    'Product Owner': [
        {
            'name': 'Glassdoor',
            'links': [
                (f'234 PO {LOCATION_CITY}', 'https://www.glassdoor.com/Job/paris-product-owner-jobs-SRCH_IL.0,5_IC2881970_KO6,19.htm')
            ]
        },
        {
            'name': 'LinkedIn',
            'links': [
                (f'1,000+ PO {LOCATION_CITY}', 'https://www.linkedin.com/jobs/product-owner-jobs-paris')
            ]
        },
        {
            'name': 'WelcomeToTheJungle',
            'links': [
                (f'Product Owner {LOCATION_CITY} (FR)', 'https://www.welcometothejungle.com/fr/jobs?query=product+owner&aroundQuery=Paris')
            ]
        }
    ]
}


def build_companies_by_role(tracker):
    # Final structure: {role_category: {company_name: company_meta}}
    companies_by_role = {}

    for company, data in tracker.items():
        # Role text from Excel (e.g. "Senior Java Developer")
        excel_role = data.get('role', '')

        # Map free-text Excel role to one of our 3 buckets
        category = map_excel_role_to_category(excel_role)
        if not category:
            # Skip rows we don’t know how to classify
            continue

        # Ensure the category exists in the dict
        companies_by_role.setdefault(category, {})

        role_link = data.get('role_link', '')

        # Create/update this company entry under that category
        companies_by_role[category][company] = {
            'industry': 'Tracker',
            'roles': excel_role or 'Various',
            'role_link': role_link,
            'experience': '',
            'links': [
                ('Search', f'https://www.google.com/search?q={company.replace(" ", "+")}+careers+paris'),
                ('LinkedIn', f'https://www.linkedin.com/company/{company.lower().replace(" ", "-")}/jobs'),
                ('WTTJ',    f'https://www.welcometothejungle.com/fr/jobs?query={company.replace(" ", "+")}'),
            ],
        }

    # Return a dict shaped like old COMPANIES_BY_ROLE, but built from Excel
    return companies_by_role



# ========================================================

def parse_hr_contacts(cell):
    """Parse HR contact cell into list of (name, url) tuples.
    Handles: plain text with hyperlink, HYPERLINK formulas, or raw text."""
    contacts = []
    if cell is None or cell.value is None:
        return contacts

    value = str(cell.value)

    # Case 1: HYPERLINK formula(s) like =HYPERLINK("url","name") & CHAR(10) & ...
    if value.startswith('='):
        # Split on CHAR(10) separator to get each part
        parts = re.split(r'\s*&\s*CHAR\(10\)\s*&\s*', value.lstrip('='))
        for part in parts:
            part = part.strip()
            hlink = re.match(r'HYPERLINK\("([^"]+)","([^"]+)"\)', part)
            if hlink:
                contacts.append((hlink.group(2), hlink.group(1)))
            else:
                # Plain text part like "alice"
                text = part.strip('"')
                if text:
                    contacts.append((text, ''))
    # Case 2: Plain text with a cell-level hyperlink
    elif cell.hyperlink:
        contacts.append((value, cell.hyperlink.target))
    # Case 3: Raw text (no links)
    elif value.strip():
        contacts.append((value.strip(), ''))

    return contacts


def read_application_tracker():
    """Read Excel tracker - called DAILY to get latest updates.
    If the file is locked (e.g. open in Excel), copies to a temp file first."""
    temp_file = None
    try:
        # Try reading directly first; if locked, copy to temp
        try:
            wb = openpyxl.load_workbook(TRACKER_FILE, data_only=False)
        except PermissionError:
            print("File is locked (Excel open?) - reading from temp copy...")
            temp_fd, temp_file = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)
            shutil.copy2(TRACKER_FILE, temp_file)
            wb = openpyxl.load_workbook(temp_file, data_only=False)
        ws = wb.active

        tracker = {}
        for i, row in enumerate(ws.iter_rows()):
            if i == 0:
                continue

            company = row[0].value
            role = row[1].value
            role_link = row[2].value if len(row) > 2 else None
            # Also check for cell-level hyperlink on role link column
            if not role_link and len(row) > 2 and row[2].hyperlink:
                role_link = row[2].hyperlink.target
            status = row[3].value
            hr_cell = row[4] if len(row) > 4 else None

            if company and 'Program/Product' not in str(company):
                company_clean = str(company).strip()
                hr_contacts = parse_hr_contacts(hr_cell)
                role_link_str = str(role_link).strip() if role_link else ''
                # Only keep if it looks like a URL
                if role_link_str and not role_link_str.startswith('http'):
                    role_link_str = ''

                if company_clean not in tracker:
                    tracker[company_clean] = {'role': role, 'role_link': role_link_str, 'status': status, 'hr_contacts': hr_contacts}
                else:
                    # Deduplicate - prioritize better status
                    current_status = str(tracker[company_clean].get('status', '')).lower()
                    new_status = str(status).lower() if status else ''
                    priority = {'review': 5, 'progress': 4, 'done': 3, 'reject': 2}
                    current_p = max([v for k, v in priority.items() if k in current_status] or [0])
                    new_p = max([v for k, v in priority.items() if k in new_status] or [0])
                    if new_p > current_p:
                        tracker[company_clean] = {'role': role, 'role_link': role_link_str, 'status': status, 'hr_contacts': hr_contacts}
                    # Merge HR contacts if current entry has none
                    if not tracker[company_clean].get('hr_contacts') and hr_contacts:
                        tracker[company_clean]['hr_contacts'] = hr_contacts
                    # Merge role_link if current entry has none
                    if not tracker[company_clean].get('role_link') and role_link_str:
                        tracker[company_clean]['role_link'] = role_link_str

        wb.close()
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)
        print(f"Tracker updated from Excel: {len(tracker)} companies")
        return tracker
    except Exception as e:
        if temp_file and os.path.exists(temp_file):
            os.remove(temp_file)
        print(f"Warning: Could not read tracker: {e}")
        return {}

def get_status_priority(status):
    """Get priority for sorting"""
    if not status:
        return 0
    status_lower = str(status).lower()
    if 'review' in status_lower or 'progress' in status_lower:
        return 1
    elif 'done' in status_lower or 'applied' in status_lower:
        return 2
    # Not available in status/role or nothing in status(immediately after applied)
    elif 'not available' in status_lower or 'nothing' in status_lower:
        return 2.5
    elif 'reject' in status_lower:
        return 3
    else:
        return 0

def get_status_compact(company, tracker):
    """Get compact status text with abbreviations"""
    if company not in tracker:
        return '<span class="s-new">⬜ NC</span>'

    status = tracker[company].get('status', '')
    status_lower = str(status).lower() if status else ''

    if 'done' in status_lower or 'applied' in status_lower:
        return '<span class="s-applied">✅ Applied</span>'
    elif 'reject' in status_lower:
        return '<span class="s-rejected">❌ Rejected</span>'
    elif 'review' in status_lower:
        return '<span class="s-review">🕐 Review</span>'
    elif 'progress' in status_lower:
        return '<span class="s-progress">🔄 Progress</span>'
    elif 'nothing' in status_lower:
        return '<span class="s-nothing">⏸️ No jobs</span>'
    else:
        return '<span class="s-new">⬜ NC</span>'

def map_excel_role_to_category(excel_role):
    """Map Excel role to our category"""
    if not excel_role:
        return None
    role_lower = str(excel_role).lower()
    # Handle 'Not available' placeholder explicitly
    if 'not available' in role_lower:
        return 'Software Developer (Java) - SENIOR/EXPERT'
    # Java Developer roles
    if any(x in role_lower for x in ['java', 'backend', 'software engineer', 'full software', 'lead software']):
        if 'backend' in role_lower or 'specialist' in role_lower:
            return 'Backend Java Developer - SENIOR'
        return 'Software Developer (Java) - SENIOR/EXPERT'

    # Product/Project Manager roles
    if any(x in role_lower for x in ['product', 'project', 'program']):
        return 'Product Owner'

    # Engineering Manager roles
    if 'engineering manager' in role_lower or 'manager' in role_lower:
        return 'Software Developer (Java) - SENIOR/EXPERT'

    return None

# ============= HOT JOBS — LinkedIn Listings =============

HOT_JOBS_HISTORY_FILE = os.path.join(os.path.dirname(__file__), 'daily_hot_jobs.json')

DEFAULT_HOT_JOB_QUERIES = {
    'Senior Java': [
        ('senior+java+developer', 'Paris, France'),
        ('senior+java+developer', 'France'),
        ('senior+software+engineer+java', 'Paris, France'),
    ],
    'Backend Java': [
        ('backend+java+developer', 'Paris, France'),
        ('lead+backend+engineer', 'France'),
    ],
    'Product Owner': [
        ('product+owner', 'Paris, France'),
        ('product+owner', 'France'),
    ],
    'Assistant Project Manager': [
        ('assistant+project+manager', 'Paris, France'),
        ('assistant+project+manager', 'France'),
        ('assistant+project+manager+java', 'Paris, France'),
        ('assistant+project+manager+java', 'France'),
    ],
}


def fetch_linkedin_jobs(keywords, location):
    """Fetch jobs from LinkedIn guest API for a single query."""
    jobs = []
    try:
        url = (
            f'https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search'
            f'?keywords={keywords}&location={location.replace(" ", "+")}&start=0'
        )
        resp = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
        if resp.status_code != 200:
            return jobs

        titles = re.findall(r'base-search-card__title[^>]*>([^<]+)<', resp.text)
        companies = re.findall(r'base-search-card__subtitle[^>]*>[^<]*<a[^>]*>([^<]+)<', resp.text)
        locations = re.findall(r'job-search-card__location[^>]*>([^<]+)<', resp.text)
        links = re.findall(r'href="(https://(?:fr|www)\.linkedin\.com/jobs/view/[^"]+)"', resp.text)

        for i in range(min(len(titles), len(companies), len(locations), len(links))):
            clean_url = unescape(links[i]).split('?')[0]
            jobs.append({
                'company': unescape(companies[i].strip()),
                'title': unescape(titles[i].strip()),
                'url': clean_url,
                'location': unescape(locations[i].strip()),
            })
    except Exception as e:
        print(f"  LinkedIn hot jobs error ({keywords}, {location}): {e}")
    return jobs


def get_hot_job_location_tier(location):
    """Return location priority: Paris(0) → France(1) → EMEA(2) → Other(3)."""
    loc = location.lower()
    if 'paris' in loc:
        return 0
    if 'france' in loc or 'île-de-france' in loc or 'ile-de-france' in loc:
        return 1
    emea = ['europe', 'emea', 'germany', 'netherlands', 'belgium', 'spain',
            'italy', 'switzerland', 'uk', 'united kingdom', 'ireland',
            'sweden', 'denmark', 'portugal', 'austria', 'poland']
    if any(e in loc for e in emea):
        return 2
    return 3


def _load_hot_jobs_file():
    """Load the full hot jobs state file."""
    try:
        with open(HOT_JOBS_HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def load_hot_jobs_current():
    """Load the current sticky hot jobs list per category."""
    return _load_hot_jobs_file().get('current_jobs', {})


def load_hot_jobs_blocklist():
    """Load the blocklist of manually removed companies."""
    return set(_load_hot_jobs_file().get('blocklist', []))


def save_hot_jobs_current(current_jobs, blocklist=None):
    """Save the current sticky hot jobs list and optional blocklist."""
    existing = _load_hot_jobs_file()
    data = {
        'last_updated': datetime.now().strftime('%Y-%m-%d'),
        'current_jobs': current_jobs,
        'blocklist': list(blocklist) if blocklist is not None else existing.get('blocklist', []),
    }
    with open(HOT_JOBS_HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2)


def _is_in_tracker(job_company, tracker_names):
    """Check if a job company matches any tracker company (substring both ways)."""
    return any(t in job_company or job_company in t for t in tracker_names)


def _is_blocklisted(job_company, job_title, blocklist):
    """Check if a specific job (company+role) is blocklisted."""
    for entry in blocklist:
        if '||' in entry:
            bl_company, bl_role = entry.split('||', 1)
            if bl_company in job_company and bl_role in job_title:
                return True
        else:
            # Legacy: company-only blocklist entry
            if entry in job_company or job_company in entry:
                return True
    return False


def fetch_hot_jobs(tracker):
    """Sticky hot jobs: keep showing the same 5 per category, only backfill gaps.

    A job is removed from the list when its company appears in the tracker.
    Only then do we fetch from LinkedIn to fill the empty slot(s).
    """
    queries = HOT_JOB_QUERIES or DEFAULT_HOT_JOB_QUERIES
    current = load_hot_jobs_current()
    blocklist = load_hot_jobs_blocklist()
    tracker_names = [name.lower().strip() for name in tracker.keys()]

    hot_jobs_by_category = {}

    for category, query_list in queries.items():
        # Start with existing sticky list for this category
        existing = current.get(category, [])

        # Remove jobs whose company is now in the tracker
        kept = []
        for job in existing:
            if _is_in_tracker(job['company'].lower().strip(), tracker_names):
                print(f"  [{category}] Removed '{job['company']}' (now in tracker)")
            else:
                kept.append(job)

        slots_needed = 5 - len(kept)

        # Only fetch from LinkedIn if we have empty slots
        if slots_needed > 0:
            print(f"  [{category}] {len(kept)} kept, need {slots_needed} more - fetching LinkedIn...")
            # Collect existing URLs/keys to avoid duplicates
            existing_keys = set(
                (j['company'].lower().strip(), j['title'].lower().strip()) for j in kept
            )
            existing_urls = set(j['url'] for j in kept)

            candidates = []
            for keywords, location in query_list:
                jobs = fetch_linkedin_jobs(keywords, location)
                print(f"    '{keywords}' in '{location}': {len(jobs)} results")
                for job in jobs:
                    if job['url'] in existing_urls:
                        continue
                    key = (job['company'].lower().strip(), job['title'].lower().strip())
                    if key in existing_keys:
                        continue
                    if _is_in_tracker(job['company'].lower().strip(), tracker_names):
                        continue
                    if _is_blocklisted(job['company'].lower().strip(), job['title'].lower().strip(), blocklist):
                        continue
                    existing_urls.add(job['url'])
                    existing_keys.add(key)
                    candidates.append(job)
                time.sleep(2)

            # Sort candidates by location tier, pick best ones to fill slots
            candidates.sort(key=lambda j: get_hot_job_location_tier(j['location']))
            kept.extend(candidates[:slots_needed])
        else:
            print(f"  [{category}] All 5 slots filled - no fetch needed")

        if kept:
            hot_jobs_by_category[category] = kept

    # Save updated sticky list
    save_hot_jobs_current(hot_jobs_by_category)

    return hot_jobs_by_category


TIER_BADGES = {0: '🏠 Paris', 1: '🇫🇷 France', 2: '🌍 EMEA', 3: '📍 Other'}


def build_hot_jobs_html(hot_jobs_by_category):
    """Build orange-themed HTML section for hot jobs."""
    if not hot_jobs_by_category:
        return ''

    total = sum(len(jobs) for jobs in hot_jobs_by_category.values())

    html = f"""
        <div style="background: linear-gradient(135deg, #e65100 0%, #ff9800 100%); color: white; padding: 8px 12px; border-radius: 6px; margin: 10px 0 5px 0;">
            <h2 style="margin: 0; font-size: 16px; color: white; border: none; padding: 0;">🔥 Hot Jobs — {total} New Listing{'s' if total != 1 else ''} Today</h2>
        </div>
"""

    for category, jobs in hot_jobs_by_category.items():
        html += f"""
        <h3 style="color: #e65100; margin: 8px 0 3px 0; font-size: 12px; padding-left: 5px;">{category} ({len(jobs)} job{'s' if len(jobs) != 1 else ''})</h3>
        <table style="border-collapse: collapse; width: 100%; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.08); font-size: 11px; margin: 3px 0;">
            <thead>
                <tr>
                    <th style="background: linear-gradient(135deg, #e65100 0%, #ff9800 100%); color: white; font-weight: bold; padding: 4px 6px; text-align: left; font-size: 10px;" width="20%">Company</th>
                    <th style="background: linear-gradient(135deg, #e65100 0%, #ff9800 100%); color: white; font-weight: bold; padding: 4px 6px; text-align: left; font-size: 10px;" width="50%">Role</th>
                    <th style="background: linear-gradient(135deg, #e65100 0%, #ff9800 100%); color: white; font-weight: bold; padding: 4px 6px; text-align: left; font-size: 10px;" width="30%">Location</th>
                </tr>
            </thead>
            <tbody>
"""
        for job in jobs:
            tier = get_hot_job_location_tier(job['location'])
            badge = TIER_BADGES.get(tier, '')
            html += f"""                <tr style="border-bottom: 1px solid #ecf0f1;">
                    <td style="padding: 4px 6px; font-weight: bold;">{job['company']}</td>
                    <td style="padding: 4px 6px;"><a href="{job['url']}" style="color: #e65100; text-decoration: underline;">{job['title']}</a></td>
                    <td style="padding: 4px 6px;"><span style="background: #fff3e0; padding: 1px 6px; border-radius: 8px; font-size: 9px;">{badge}</span> {job['location']}</td>
                </tr>
"""
        html += """            </tbody>
        </table>
"""

    return html


def create_job_report():
    """Generate ULTRA COMPACT job report"""
    tracker = read_application_tracker()

    # Fetch hot jobs from LinkedIn (filtered against tracker + history)
    print("Fetching hot jobs from LinkedIn...")
    hot_jobs_by_category = fetch_hot_jobs(tracker)
    hot_jobs_total = sum(len(jobs) for jobs in hot_jobs_by_category.values())
    print(f"Hot jobs: {hot_jobs_total} listings across {len(hot_jobs_by_category)} categories")
    hot_jobs_html = build_hot_jobs_html(hot_jobs_by_category)

    # Merge tracker companies into COMPANIES_BY_ROLE
   # companies_merged = {}
   # for role_name in COMPANIES_BY_ROLE.keys():
   #     companies_merged[role_name] = dict(COMPANIES_BY_ROLE[role_name])
   
    companies_merged = build_companies_by_role(tracker)
    # Add all companies from Excel tracker
    for company, data in tracker.items():
        excel_role = data.get('role', '')
        category = map_excel_role_to_category(excel_role)

        if category and category in companies_merged:
            # Add if not already in the role
            if company not in companies_merged[category]:
                companies_merged[category][company] = {
                    'industry': 'Tracker',
                    'roles': excel_role if excel_role else 'Various',
                    'role_link': data.get('role_link', ''),
                    'experience': '',
                    'links': [
                        ('Search', f'https://www.google.com/search?q={company.replace(" ", "+")}+careers+paris'),
                        ('LinkedIn', f'https://www.linkedin.com/company/{company.lower().replace(" ", "-")}/jobs')
                    ]
                }

    # Count total stats
    all_job_search_companies = set()
    for role_companies in companies_merged.values():
        all_job_search_companies.update(role_companies.keys())

    all_companies = set(list(tracker.keys()) + list(all_job_search_companies))
    not_contacted = sum(1 for c in all_job_search_companies if c not in tracker)
    applied = len([c for c, d in tracker.items() if d.get('status') and 'done' in str(d.get('status')).lower()])
    under_review = len([c for c, d in tracker.items() if d.get('status') and 'review' in str(d.get('status')).lower()])
    rejected = len([c for c, d in tracker.items() if d.get('status') and 'reject' in str(d.get('status')).lower()])
    nothing = len([c for c, d in tracker.items()  if d.get('status') and 'nothing' in str(d.get('status')).lower()])
    report = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 11px; line-height: 1.2; color: #2c3e50; max-width: 1600px; margin: 0; padding: 8px; background: #f8f9fa; }}
            h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 5px; margin: 5px 0; font-size: 18px; }}
            h2 {{ color: #3498db; margin: 15px 0 5px 0; font-size: 14px; border-left: 3px solid #3498db; padding-left: 8px; }}
            h3 {{ color: #27ae60; margin: 10px 0 3px 0; font-size: 12px; padding-left: 5px; }}
            .header-info {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 6px 10px; border-radius: 4px; margin-bottom: 8px; font-size: 11px; }}
            .header-info p {{ margin: 2px 0; }}
            .experience-badge {{ background: #e74c3c; color: white; padding: 2px 6px; border-radius: 3px; font-size: 9px; font-weight: bold; margin-left: 8px; }}
            table {{ border-collapse: collapse; width: 100%; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.08); font-size: 11px; margin: 5px 0; }}
            th {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; font-weight: bold; padding: 5px 6px; text-align: left; font-size: 10px; }}
            td {{ padding: 4px 6px; border-bottom: 1px solid #ecf0f1; vertical-align: top; }}
            tr:hover {{ background-color: #f8f9fa; }}
            tr.section-header {{ background: #e8f5e9; font-weight: bold; }}
            tr.section-header td {{ padding: 4px 6px; font-size: 11px; color: #2c3e50; }}
            .company-name {{ font-weight: bold; color: #2c3e50; font-size: 12px; }}
            .new-badge {{ background: #27ae60; color: white; padding: 1px 5px; border-radius: 8px; font-size: 9px; font-weight: bold; margin-left: 5px; }}
            .links a {{ color: #3498db; text-decoration: none; font-size: 10px; display: inline-block; margin: 1px 5px 1px 0; }}
            .links a:before {{ content: "→ "; }}
            .links a:hover {{ text-decoration: underline; }}
            .s-applied {{ color: #27ae60; font-weight: bold; font-size: 10px; }}
            .s-rejected {{ color: #e74c3c; font-weight: bold; font-size: 10px; }}
            .s-review {{ color: #f39c12; font-weight: bold; font-size: 10px; }}
            .s-new {{ color: #3498db; font-weight: bold; font-size: 10px; }}
            .s-nothing {{ color: #6c757d; font-weight: bold; font-size: 10px; }}
            .s-progress {{ color: #17a2b8; font-weight: bold; font-size: 10px; }}
            .footer {{ text-align: center; color: #7f8c8d; font-size: 10px; margin-top: 12px; padding-top: 8px; border-top: 2px solid #ecf0f1; }}
            .summary {{ background: #d4edda; padding: 5px 8px; border-radius: 3px; margin: 5px 0; font-size: 11px; }}
            .note {{ font-size: 10px; color: #6c757d; }}
            .exp-note {{ font-size: 9px; color: #e74c3c; font-weight: bold; }}
            .hr-contact a {{ color: #8e44ad; text-decoration: none; font-size: 10px; display: block; }}
            .hr-contact a:hover {{ text-decoration: underline; }}
        </style>
    </head>
    <body>
        <h1>🎯 Senior Jobs (8+ Years)<span class="experience-badge">COMPACT</span></h1>

        <div class="header-info">
            <p>📅 {datetime.now().strftime('%Y-%m-%d %H:%M')} | 📊 {len(all_companies)} Companies | 🔍 {LOCATION_CITY}/{LOCATION_COUNTRY} | Senior Java (8+ yrs) | ⬜ NC: {not_contacted} | ✅ Applied: {applied} |⏸️ No Jobs Available: {nothing}|🕐 Review: {under_review} | ❌ Rejected: {rejected}</p>
             <p>NOTE: ALL JOBS WHICH ARE NOT AVAILABLE OR NOTHING IN STATUS TRACKER LIST.XLSX goes as  No Jobs Available under Senior Java (8+ yrs)  </p>
        </div>
"""

    # Inject hot jobs section
    report += hot_jobs_html

    # Generate tables for each role
    for role_name, companies_dict in companies_merged.items():
        report += f"""
        <h2>{role_name}</h2>
        <table>
            <thead>
                <tr>
                    <th width="14%">Company</th>
                    <th width="8%">Status</th>
                    <th width="22%">Role</th>
                    <th width="16%">HR Contact</th>
                    <th width="40%">Links</th>
                </tr>
            </thead>
            <tbody>
"""

        # Sort companies by status
        companies_sorted = sorted(companies_dict.items(),
                                 key=lambda x: (get_status_priority(tracker.get(x[0], {}).get('status')), x[0]))

        current_section = None
        for company_name, company_info in companies_sorted:
            status = tracker.get(company_name, {}).get('status', '')
            priority = get_status_priority(status)

            # Section headers (more compact)
            section_name = None
            if priority == 0 and current_section != 'not_contacted':
                section_name = '⬜ NOT CONTACTED'
                current_section = 'not_contacted'
            elif priority == 1 and current_section != 'in_progress':
                section_name = '🕐 UNDER REVIEW / IN PROGRESS'
                current_section = 'in_progress'
            elif priority == 2 and current_section != 'applied':
                section_name = '✅ APPLIED'
                current_section = 'applied'
            elif priority == 2.5 and current_section != 'nothing':
                section_name = '⏸️ No Jobs Available'
                current_section = 'nothing'
            elif priority == 3 and current_section != 'rejected':
                section_name = '❌ REJECTED'
                current_section = 'rejected'

            if section_name:
                report += f'                <tr class="section-header"><td colspan="5">{section_name}</td></tr>\n'

            # Company row
            is_new = company_name not in tracker
            industry = company_info.get('industry', '')
            roles = company_info.get('roles', '')
            role_link = company_info.get('role_link', '') or tracker.get(company_name, {}).get('role_link', '')
            links = company_info.get('links', [])

            # Make role text clickable if we have a role link
            if role_link:
                role_html = f'<a href="{role_link}" style="color: #2c3e50; text-decoration: underline;">{roles}</a>'
            else:
                role_html = roles

            # Get HR contacts for this company
            hr_contacts = tracker.get(company_name, {}).get('hr_contacts', [])
            hr_html = ''
            if hr_contacts:
                for name, url in hr_contacts:
                    if url:
                        hr_html += f'<a href="{url}">{name}</a>'
                    else:
                        hr_html += f'{name}'
            else:
                hr_html = '<span class="note">-</span>'

            report += f"""                <tr>
                    <td>
                        <div class="company-name">{company_name}{' <span class="new-badge">NEW</span>' if is_new else ''}</div>
                        <div class="note">{industry}</div>
                    </td>
                    <td>{get_status_compact(company_name, tracker)}</td>
                    <td>
                        <div style="font-size: 11px;">{role_html}</div>
                    </td>
                    <td class="hr-contact">{hr_html}</td>
                    <td class="links">
"""
            for link_text, link_url in links:
                report += f'<a href="{link_url}">{link_text}</a> '

            report += """
                    </td>
                </tr>
"""

        report += """            </tbody>
        </table>
"""

        # Add platform aggregators under each role (compact)
        if role_name in PLATFORM_AGGREGATORS:
            report += f"""
        <h3>🔍 More {role_name.split(' - ')[0]}:</h3>
        <table>
            <tbody>
"""
            for platform in PLATFORM_AGGREGATORS[role_name]:
                report += f"""                <tr>
                    <td width="16%"><strong>{platform['name']}</strong></td>
                    <td class="links">
"""
                for link_text, link_url in platform['links']:
                    report += f'<a href="{link_url}">{link_text}</a> '

                report += """</td>
                </tr>
"""

            report += """            </tbody>
        </table>
"""

    report += f"""
        <div class="footer">
            <p>📊 {len(all_companies)} companies | ⬜ {not_contacted} NC | ✅ {applied} Applied | 🕐 {under_review} Review | ❌ {rejected} Rejected | 🔄 Next: Tomorrow 11:00 CET</p>
        </div>
    </body>
    </html>
    """

    return report

def send_email(html_content):
    """Send email"""
    try:
        message = MIMEMultipart("alternative")
        message["Subject"] = f"Senior Jobs COMPACT - {datetime.now().strftime('%Y-%m-%d')}"
        message["From"] = EMAIL_CONFIG['sender_email']
        message["To"] = EMAIL_CONFIG['recipient_email']

        html_part = MIMEText(html_content, "html")
        message.attach(html_part)

        with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
            server.send_message(message)

        print(f"SUCCESS: Compact email sent")
        return True
    except Exception as e:
        print(f"ERROR: {str(e)}")
        return False

def main():
    """Main - Reads Excel tracker DAILY (skips Friday)"""
    if datetime.now().weekday() == 4:  # 4 = Friday
        print("Friday - skipping email send")
        return

    print("Reading tracker from Excel (daily update)...")
    tracker = read_application_tracker()
    print(f"Found {len(tracker)} companies in tracker")

    print("Generating ULTRA COMPACT report...")
    report_html = create_job_report()

    print("Sending email...")
    send_email(report_html)

    # Generate outreach drafts for applied companies
    from outreach_drafter import run_outreach
    run_outreach()

    # Generate tailored resumes for applied companies
    from resume_tailor import run_tailor
    run_tailor()

def run_hot_jobs_only():
    """Standalone hot jobs check - prints to console, no email.

    Flags:
        --refresh           Clear all categories and re-fetch
        --refresh "Cat"     Clear only that category and re-fetch
    """
    import sys
    print("=== Hot Jobs Check ===")
    tracker = read_application_tracker()
    print(f"Tracker: {len(tracker)} companies\n")

    # Handle --remove flag: drop a specific job (company + role) and blocklist it
    # Usage: --remove "Company" "Role title"
    if '--remove' in sys.argv:
        remove_idx = sys.argv.index('--remove')
        remaining = [a for a in sys.argv[remove_idx + 1:] if not a.startswith('--')]
        if len(remaining) >= 2:
            remove_company = remaining[0].lower().strip()
            remove_role = remaining[1].lower().strip()
            current = load_hot_jobs_current()
            blocklist = load_hot_jobs_blocklist()
            removed = False
            for cat, jobs in current.items():
                before = len(jobs)
                current[cat] = [
                    j for j in jobs
                    if not (remove_company in j['company'].lower() and remove_role in j['title'].lower())
                ]
                if len(current[cat]) < before:
                    print(f"Removed '{remove_company}' / '{remove_role}' from {cat}")
                    removed = True
            blocklist.add(f"{remove_company}||{remove_role}")
            save_hot_jobs_current(current, blocklist)
            if removed:
                print(f"Blocklisted - this job won't reappear")
            else:
                print(f"Job not found in current list but blocklisted for future")
            print()
        elif len(remaining) == 1:
            print("Usage: --remove \"Company\" \"Role title\"")
            print("Both company and role are required to avoid removing other listings")
            print()
            return

    # Handle --refresh flag
    if '--refresh' in sys.argv:
        refresh_idx = sys.argv.index('--refresh')
        # Check if a specific category was given after --refresh
        refresh_cat = None
        if refresh_idx + 1 < len(sys.argv) and not sys.argv[refresh_idx + 1].startswith('--'):
            refresh_cat = sys.argv[refresh_idx + 1]

        current = load_hot_jobs_current()
        if refresh_cat:
            if refresh_cat in current:
                del current[refresh_cat]
                print(f"Cleared '{refresh_cat}' - will re-fetch\n")
            else:
                print(f"Category '{refresh_cat}' not found. Available: {', '.join(current.keys())}\n")
        else:
            current = {}
            print("Cleared all categories - will re-fetch\n")
        save_hot_jobs_current(current)

    hot_jobs_by_category = fetch_hot_jobs(tracker)
    total = sum(len(jobs) for jobs in hot_jobs_by_category.values())

    if not hot_jobs_by_category:
        print("No hot jobs found.")
        return

    tier_labels = {0: '[Paris]', 1: '[France]', 2: '[EMEA]', 3: '[Other]'}

    print(f"\n{'='*60}")
    print(f"  {total} Hot Jobs across {len(hot_jobs_by_category)} categories")
    print(f"{'='*60}\n")

    for category, jobs in hot_jobs_by_category.items():
        print(f"  {category} ({len(jobs)})")
        print(f"  {'-'*40}")
        for job in jobs:
            tier = get_hot_job_location_tier(job['location'])
            label = tier_labels.get(tier, '')
            print(f"    {job['company']}")
            print(f"      {job['title']}")
            print(f"      {label} {job['location']}")
            print(f"      {job['url']}")
            print()


if __name__ == "__main__":
    import sys
    if '--hot-jobs' in sys.argv:
        run_hot_jobs_only()
    else:
        main()
