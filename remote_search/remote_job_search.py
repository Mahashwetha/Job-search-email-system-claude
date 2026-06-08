"""
Remote Job Search — fetches remote-only listings from free APIs/feeds,
filters for EMEA-compatible roles matching user profile,
and sends a styled HTML email every 2 days.

Sources: RemoteOK, Remotive, We Work Remotely, Jobicy, LinkedIn (France/Global),
Bluedoor (ATS-aggregated postings, EMEA description-verified).
"""

import sys
import os
import re
import json
import time
import shutil
import smtplib
import tempfile
import requests
import xml.etree.ElementTree as ET
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from email.utils import parsedate_to_datetime
from html import unescape

# Fit scorer (optional — disabled gracefully if unavailable)
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
try:
    from fit_scorer import score_fit_batch, fit_badge_html, FIT_SCORE_ENABLED
    _FIT_AVAILABLE = FIT_SCORE_ENABLED
except ImportError:
    _FIT_AVAILABLE = False
    def fit_badge_html(_): return ''

# Add parent directory to path so we can import config
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

try:
    from config import EMAIL_CONFIG
except ImportError:
    print("ERROR: config.py not found!")
    print("Please copy config.template.py to config.py and fill in your details.")
    exit(1)

# Import optional remote search config from config.py, with defaults
try:
    from config import REMOTE_ROLE_KEYWORDS, REMOTE_LOCATION_INCLUDE, REMOTE_LOCATION_EXCLUDE
except ImportError:
    REMOTE_ROLE_KEYWORDS = None
    REMOTE_LOCATION_INCLUDE = None
    REMOTE_LOCATION_EXCLUDE = None

# Bluedoor is an optional extra source (ATS-aggregated postings). Off in the
# template (most users won't need it); defaults to True if the flag is absent so
# existing configs keep working. Set BLUEDOOR_ENABLED = False in config.py to skip.
try:
    from config import BLUEDOOR_ENABLED
except ImportError:
    BLUEDOOR_ENABLED = True

# ============= FILTERING CONFIG =============
# Override these in config.py to customize for your profile and location

ROLE_KEYWORDS = REMOTE_ROLE_KEYWORDS or [
    # Primary: Java/backend (highest relevance)
    'java', 'backend', 'back-end', 'back end',
    'software engineer', 'senior software', 'tech lead',
    'api engineer',
    # AI roles (not ML/data science prerequisite)
    'ai engineer', 'genai', 'llm engineer', 'llm',
    'generative ai', 'ai platform engineer', 'foundation model',
]

# Python is secondary — only include Python roles if they ALSO mention Java/backend
PYTHON_SECONDARY_KEYWORDS = ['python']
JAVA_BACKEND_SIGNALS = ['java', 'backend', 'back-end', 'back end', 'jvm', 'spring', 'microservices']

# Sources that are inherently EU-focused (jobs from these pass without explicit EMEA location)
EU_FOCUSED_SOURCES = []

# Sources where "Remote" without region is common — allow if no US indicators found
RELAXED_LOCATION_SOURCES = ['Jobicy']

# Sources searched with EMEA keywords — bypass LOCATION_EXCLUDE (already pre-filtered)
EMEA_SEARCHED_SOURCES = ['LinkedIn Global']

# ── Rejected Remote List ──
# Loaded from rejected_remote.json — use reject_remote.py to add entries.
# Each entry is [company_substring, title_substring] — both lowercased.
# To review a job again, run: python reject_remote.py --remove "company" "title"
_REJECTED_FILE = os.path.join(os.path.dirname(__file__), 'rejected_remote.json')

def _load_rejected():
    try:
        with open(_REJECTED_FILE, 'r', encoding='utf-8') as f:
            return [tuple(e) for e in json.load(f)]
    except (FileNotFoundError, json.JSONDecodeError):
        return []

REJECTED_REMOTE_LIST = _load_rejected()

# Roles to exclude even if they match keywords above
ROLE_EXCLUDE = [
    'data scientist', 'machine learning engineer', 'ml engineer',
    'data engineer', 'analytics engineer', 'research scientist',
    'frontend', 'front-end', 'front end', 'ios ', 'android ',
    'react native', 'flutter',
    'web developer', 'wordpress', 'php developer', 'ruby',
    'salesforce', '.net', 'dotnet', 'c# ', 'node.js',
    'hubspot', 'webflow', 'figma', 'shopify',
    'account executive', 'sales ', 'marketing',
    'full stack', 'full-stack', 'fullstack',
    'network engineer', 'voip', 'linux admin',
    'consultant', 'werkstudent', 'intern ',
    'new grad', 'graduate engineer', 'entry level', 'junior ',
    'guatemala', 'latin america', 'latam',
    'ror ', 'rails developer', 'ruby on rails',
    'web development manager', 'manager, web',
    'freelance', 'dataops', 'data ops',
    'network automation',
    'devops', 'dev ops', 'cloud engineer', 'sre',
    'site reliability', 'infrastructure engineer',
    'platform engineer', 'systems engineer',
    'prompt engineer', 'c++',
]

# Location priority tiers for sorting (lower = shown first)
LOCATION_PRIORITY = [
    # Tier 0: Paris
    (['paris'], 0),
    # Tier 1: France
    (['france'], 1),
    # Tier 2: EMEA / CET timezone countries
    (['emea', 'europe', 'eu', 'germany', 'netherlands', 'belgium',
      'spain', 'italy', 'switzerland', 'austria', 'poland', 'czech',
      'sweden', 'norway', 'denmark', 'portugal', 'ireland',
      'cet', 'cest', 'central european'], 2),
    # Tier 3: UK (GMT+0/+1, close to CET)
    (['uk', 'united kingdom', 'london', 'britain'], 3),
    # Tier 4: Worldwide/anywhere/global
    (['worldwide', 'anywhere', 'global'], 4),
]

LOCATION_INCLUDE = REMOTE_LOCATION_INCLUDE or [
    'worldwide', 'anywhere', 'emea', 'europe', 'eu', 'france',
    'paris', 'global', 'uk', 'germany', 'netherlands',
    'belgium', 'spain', 'italy', 'switzerland', 'austria',
    'sweden', 'norway', 'denmark', 'portugal', 'ireland',
    'poland', 'czech', 'united kingdom', 'london',
    'romania', 'hungary', 'greece', 'finland', 'croatia',
    'luxembourg', 'berlin', 'amsterdam', 'barcelona', 'munich',
    'dublin', 'lisbon', 'warsaw', 'prague', 'vienna', 'brussels',
    'gmt', 'cet', 'cest', 'central european', 'utc+1', 'utc+2',
    # Note: bare 'remote' excluded to avoid US-default listings
]

LOCATION_EXCLUDE = REMOTE_LOCATION_EXCLUDE or [
    'us only', 'us timezone', 'americas only', 'usa only',
    'us-based', 'est/pst', 'canada only', 'canada', 'north america',
    'us or canada', 'usa/canada', 'na only',
    'new york', 'san francisco', 'los angeles', 'seattle', 'chicago',
    'austin', 'boston', 'denver', 'miami', 'toronto', 'vancouver',
    'asia only', 'apac only', 'india only', 'latam only',
    'united states', 'seattle, wa', 'washington, dc',
]

# ============= API FETCHERS =============

def fetch_remoteok():
    """Fetch jobs from RemoteOK API."""
    jobs = []
    try:
        resp = requests.get(
            'https://remoteok.com/api',
            headers={'User-Agent': 'RemoteJobSearch/1.0'},
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()
        for item in data[1:]:
            jobs.append({
                'company': item.get('company', ''),
                'title': item.get('position', ''),
                'url': item.get('url', ''),
                'source': 'RemoteOK',
                'location': item.get('location', 'Remote'),
                'tags': ', '.join(item.get('tags', [])),
                'posted_date': item.get('date', '')[:10],
            })
        print(f"  RemoteOK: {len(jobs)} jobs fetched")
    except Exception as e:
        print(f"  RemoteOK error: {e}")
    return jobs


def fetch_remotive():
    """Fetch jobs from Remotive API."""
    jobs = []
    try:
        resp = requests.get(
            'https://remotive.com/api/remote-jobs',
            timeout=15,
        )
        resp.raise_for_status()
        data = resp.json()
        for item in data.get('jobs', []):
            jobs.append({
                'company': item.get('company_name', ''),
                'title': item.get('title', ''),
                'url': item.get('url', ''),
                'source': 'Remotive',
                'location': item.get('candidate_required_location', 'Remote'),
                'tags': item.get('category', ''),
                'posted_date': item.get('publication_date', '')[:10],
            })
        print(f"  Remotive: {len(jobs)} jobs fetched")
    except Exception as e:
        print(f"  Remotive error: {e}")
    return jobs


def fetch_arbeitnow():
    """Fetch jobs from Arbeitnow API (EU-focused), multiple pages."""
    jobs = []
    try:
        for page in range(1, 4):  # Fetch up to 3 pages
            resp = requests.get(
                f'https://www.arbeitnow.com/api/job-board-api?page={page}',
                timeout=15,
            )
            resp.raise_for_status()
            data = resp.json()
            page_data = data.get('data', [])
            if not page_data:
                break
            for item in page_data:
                if not item.get('remote', False):
                    continue
                created_at = item.get('created_at', '')
                if isinstance(created_at, int):
                    posted = datetime.fromtimestamp(created_at).strftime('%Y-%m-%d')
                else:
                    posted = str(created_at)[:10]
                jobs.append({
                    'company': item.get('company_name', ''),
                    'title': item.get('title', ''),
                    'url': item.get('url', ''),
                    'source': 'Arbeitnow',
                    'location': item.get('location', 'Remote'),
                    'tags': ', '.join(item.get('tags', [])),
                    'posted_date': posted,
                })
        print(f"  Arbeitnow: {len(jobs)} remote jobs fetched")
    except Exception as e:
        print(f"  Arbeitnow error: {e}")
    return jobs


def fetch_jobicy():
    """Fetch jobs from Jobicy RSS feed (remote tech jobs)."""
    jobs = []
    try:
        resp = requests.get(
            'https://jobicy.com/feed/newjobs',
            headers={'User-Agent': 'RemoteJobSearch/1.0'},
            timeout=15,
        )
        resp.raise_for_status()
        root = ET.fromstring(resp.text)

        for item in root.findall('./jobs/job'):
            pub_date = item.findtext('pubdate', '')
            posted = ''
            if pub_date:
                try:
                    # Jobicy uses DD.MM.YYYY format
                    posted = datetime.strptime(pub_date.strip(), '%d.%m.%Y').strftime('%Y-%m-%d')
                except Exception:
                    posted = pub_date[:10]

            jobs.append({
                'company': unescape(item.findtext('company', '')),
                'title': unescape(item.findtext('name', '')),
                'url': item.findtext('link', ''),
                'source': 'Jobicy',
                'location': item.findtext('region', 'Remote'),
                'tags': item.findtext('jobtype', ''),
                'posted_date': posted,
            })
        print(f"  Jobicy: {len(jobs)} jobs fetched")
    except Exception as e:
        print(f"  Jobicy error: {e}")
    return jobs


def fetch_weworkremotely():
    """Fetch jobs from We Work Remotely RSS feeds (programming + devops)."""
    jobs = []
    feeds = [
        'https://weworkremotely.com/categories/remote-back-end-programming-jobs.rss',
        'https://weworkremotely.com/categories/remote-programming-jobs.rss',
    ]
    seen_guids = set()
    for feed_url in feeds:
        try:
            resp = requests.get(
                feed_url,
                headers={'User-Agent': 'RemoteJobSearch/1.0'},
                timeout=15,
            )
            resp.raise_for_status()
            root = ET.fromstring(resp.text)

            for item in root.findall('.//item'):
                guid = item.findtext('guid', '')
                if guid in seen_guids:
                    continue
                seen_guids.add(guid)

                # Title format: "Company: Job Title"
                raw_title = item.findtext('title', '')
                if ':' in raw_title:
                    company, title = raw_title.split(':', 1)
                    company = company.strip()
                    title = title.strip()
                else:
                    company = ''
                    title = raw_title.strip()

                region = item.findtext('region', 'Remote')
                country = item.findtext('country', '')
                location = region
                if country and country not in region:
                    location = f"{region}, {country}"

                pub_date = item.findtext('pubDate', '')
                posted = ''
                if pub_date:
                    try:
                        posted = parsedate_to_datetime(pub_date).strftime('%Y-%m-%d')
                    except Exception:
                        posted = pub_date[:10]

                jobs.append({
                    'company': company,
                    'title': title,
                    'url': guid or item.findtext('link', ''),
                    'source': 'WWR',
                    'location': location,
                    'tags': item.findtext('category', ''),
                    'posted_date': posted,
                })
        except Exception as e:
            print(f"  WWR error ({feed_url.split('/')[-1]}): {e}")

    print(f"  WeWorkRemotely: {len(jobs)} jobs fetched")
    return jobs


def fetch_linkedin_france():
    """Fetch backend/Java jobs in France from LinkedIn's public guest API."""
    jobs = []
    queries = [
        'java+backend',
        'backend+engineer',
        'senior+software+engineer+java',
        'genai+engineer',
        'llm+engineer',
        'ai+engineer+backend',
    ]
    seen_urls = set()
    for query in queries:
        try:
            # f_WT=2 filters to remote-only positions
            url = f'https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search?keywords={query}&location=France&f_WT=2&start=0'
            resp = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
            if resp.status_code != 200:
                continue

            titles = re.findall(r'base-search-card__title[^>]*>([^<]+)<', resp.text)
            companies = re.findall(r'base-search-card__subtitle[^>]*>[^<]*<a[^>]*>([^<]+)<', resp.text)
            locations = re.findall(r'job-search-card__location[^>]*>([^<]+)<', resp.text)
            links = re.findall(r'href="(https://fr\.linkedin\.com/jobs/view/[^"]+)"', resp.text)

            for i in range(min(len(titles), len(companies), len(locations), len(links))):
                clean_url = unescape(links[i]).split('?')[0]
                if clean_url in seen_urls:
                    continue
                seen_urls.add(clean_url)

                title_text = titles[i].strip()

                jobs.append({
                    'company': companies[i].strip(),
                    'title': title_text,
                    'url': clean_url,
                    'source': 'LinkedIn FR',
                    'location': locations[i].strip(),
                    'tags': '',
                    'posted_date': datetime.now().strftime('%Y-%m-%d'),
                })
        except Exception as e:
            print(f"  LinkedIn FR error ({query}): {e}")

    print(f"  LinkedIn France: {len(jobs)} jobs fetched")
    return jobs


def _check_emea_timezone_in_description(job_id):
    """Fetch LinkedIn job description and check for explicit EMEA timezone compatibility.

    Returns: 'emea' | 'us_only' | 'unknown'
    - emea: explicit EMEA/Europe/flexible timezone signals found
    - us_only: US-only timezone signals found → reject
    - unknown: no timezone info found → reject (too risky)
    """
    try:
        url = f'https://www.linkedin.com/jobs-guest/jobs/api/jobPosting/{job_id}'
        resp = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
        if resp.status_code != 200:
            return 'unknown'

        match = re.search(r'show-more-less-html__markup[^>]*>(.*?)</div', resp.text, re.DOTALL)
        if not match:
            return 'unknown'

        text = re.sub(r'<[^>]+>', ' ', match.group(1)).lower()
        text = ' '.join(text.split())

        # Explicit EMEA/flexible timezone signals (employee location, not company market)
        emea_signals = [
            'emea', 'europe timezone', 'european timezone',
            'cet', 'cest', 'utc+1', 'utc+2', 'gmt+1', 'gmt+2',
            'work from anywhere', 'work from any', 'anywhere in the world',
            'any timezone', 'flexible timezone', 'flexible time zone',
            'all timezones', 'all time zones', 'open to all locations',
            'location agnostic', 'location-agnostic', 'fully distributed',
        ]
        # US/North America only signals
        us_only_signals = [
            'us timezone', 'us time zone', 'must be in the us', 'must be us',
            'must reside in', 'must be located in', 'must be based in the us',
            'north america only', 'us-based only', 'usa only',
            'eastern time zone', 'pacific time zone', 'mountain time zone',
            'central time zone', 'est timezone', 'pst timezone',
            'eastern or pacific', 'et/pt', 'et or pt',
        ]

        if any(s in text for s in us_only_signals):
            return 'us_only'
        if any(s in text for s in emea_signals):
            return 'emea'
        return 'unknown'
    except Exception:
        return 'unknown'


def fetch_linkedin_global():
    """Fetch backend/Java jobs in India, Boston, New York — verified EMEA timezone compatible."""
    candidates = []
    searches = [
        # (keywords, location)
        ('java+backend+EMEA', 'India'),
        ('software+engineer+EMEA', 'India'),
        ('backend+engineer+EMEA', 'India'),
        ('java+backend+EMEA', 'Boston, MA'),
        ('software+engineer+EMEA', 'Boston, MA'),
        ('java+backend+EMEA', 'New York, NY'),
        ('software+engineer+EMEA', 'New York, NY'),
        ('backend+engineer+global+remote', 'India'),
        ('backend+engineer+global+remote', 'New York, NY'),
        ('backend+engineer+global+remote', 'Boston, MA'),
        ('genai+engineer+EMEA', 'India'),
        ('llm+engineer+EMEA', 'India'),
        ('ai+engineer+EMEA+remote', 'New York, NY'),
        ('genai+engineer+global+remote', 'New York, NY'),
    ]
    seen_urls = set()
    for query, location in searches:
        try:
            url = (
                f'https://www.linkedin.com/jobs-guest/jobs/api/seeMoreJobPostings/search'
                f'?keywords={query}&location={location.replace(" ", "+")}&f_WT=2&start=0'
            )
            resp = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
            if resp.status_code != 200:
                continue

            titles = re.findall(r'base-search-card__title[^>]*>([^<]+)<', resp.text)
            companies = re.findall(r'base-search-card__subtitle[^>]*>[^<]*<a[^>]*>([^<]+)<', resp.text)
            locations = re.findall(r'job-search-card__location[^>]*>([^<]+)<', resp.text)
            links = re.findall(r'href="(https://(?:fr|www|in)\.linkedin\.com/jobs/view/[^"]+)"', resp.text)

            for i in range(min(len(titles), len(companies), len(locations), len(links))):
                clean_url = unescape(links[i]).split('?')[0]
                if clean_url in seen_urls:
                    continue
                seen_urls.add(clean_url)
                candidates.append({
                    'company': companies[i].strip(),
                    'title': titles[i].strip(),
                    'url': clean_url,
                    'source': 'LinkedIn Global',
                    'location': locations[i].strip(),
                    'tags': 'EMEA-verified',
                    'posted_date': datetime.now().strftime('%Y-%m-%d'),
                })
        except Exception as e:
            print(f"  LinkedIn Global error ({query}, {location}): {e}")

    # Verify EMEA timezone compatibility by checking each job description
    jobs = []
    print(f"  LinkedIn Global: {len(candidates)} candidates — verifying EMEA timezone...")
    for job in candidates:
        job_id_match = re.search(r'/jobs/view/[^/]+-(\d+)$', job['url'])
        if not job_id_match:
            continue
        job_id = job_id_match.group(1)
        result = _check_emea_timezone_in_description(job_id)
        if result == 'emea':
            jobs.append(job)
        time.sleep(1)  # rate limit

    print(f"  LinkedIn Global (India/Boston/NY): {len(jobs)} EMEA-compatible jobs (from {len(candidates)} candidates)")

    print(f"  LinkedIn Global (India/Boston/NY): {len(jobs)} jobs fetched")
    return jobs


# ============= BLUEDOOR (ATS-aggregated postings) =============
# Free public API (no auth) aggregating Greenhouse/Lever/Ashby/Workday/+27 more.
# Docs: https://bluedoor.sh/apis/job-postings/docs/
BLUEDOOR_BASE = 'https://api.bluedoor.sh/job-postings/v1'

# EMEA countries to pull remote jobs for — guarantees the role is workable from
# the EMEA region (a remote role scoped to these countries is EMEA-compatible).
BLUEDOOR_EMEA_COUNTRIES = [
    'France', 'United Kingdom', 'Germany', 'Netherlands', 'Ireland',
    'Spain', 'Portugal', 'Poland', 'Belgium', 'Sweden',
]

# Cache org_id -> company display name to avoid refetching the same org.
_BLUEDOOR_ORG_CACHE = {}


def _bluedoor_company_name(job):
    """Resolve a company name for a bluedoor job (org display_name, cached;
    falls back to parsing the apply/source URL host)."""
    org_id = job.get('org_id')
    if org_id and org_id in _BLUEDOOR_ORG_CACHE:
        return _BLUEDOOR_ORG_CACHE[org_id]

    name = ''
    if org_id:
        try:
            resp = requests.get(f'{BLUEDOOR_BASE}/orgs/{org_id}',
                                headers={'User-Agent': 'RemoteJobSearch/1.0'}, timeout=12)
            if resp.ok:
                data = resp.json().get('data', {})
                name = (data.get('display_name') or data.get('canonical_name') or '').strip()
        except Exception:
            name = ''
        _BLUEDOOR_ORG_CACHE[org_id] = name  # cache even empty to avoid retries

    if not name:
        # Fallback: derive from URL host subdomain (e.g. acme.bamboohr.com -> Acme)
        url = job.get('apply_url') or job.get('source_url') or ''
        m = re.search(r'https?://([^./]+)\.', url)
        if m and m.group(1) not in ('www', 'jobs', 'boards', 'careers', 'apply'):
            name = m.group(1).replace('-', ' ').title()
    return name or 'Unknown'


def fetch_bluedoor():
    """Fetch EMEA-compatible remote jobs from the bluedoor public ATS API.

    Pulls remote roles scoped to EMEA countries (so a remote job is workable
    from the EMEA region), normalizes them to the standard job dict shape, and
    lets the existing filter_jobs() apply role + US-exclusion screening on top.
    """
    from datetime import timedelta
    jobs = []
    posted_after = (datetime.now() - timedelta(days=45)).strftime('%Y-%m-%d')

    for country in BLUEDOOR_EMEA_COUNTRIES:
        cursor = None
        for _page in range(2):  # up to 2 pages per country
            try:
                params = {
                    'country': country,
                    'workplace_type': 'remote',
                    'active': 'true',
                    'posted_after': posted_after,
                    'limit': 100,
                }
                if cursor:
                    params['cursor'] = cursor
                resp = requests.get(f'{BLUEDOOR_BASE}/jobs/search', params=params,
                                    headers={'User-Agent': 'RemoteJobSearch/1.0'}, timeout=20)
                resp.raise_for_status()
                payload = resp.json()
            except Exception as e:
                print(f"  Bluedoor error ({country}): {e}")
                break

            for item in payload.get('data', []):
                # Build a real location string so filter_jobs geo screening still runs
                loc_parts = [item.get('city'), item.get('region'), item.get('country')]
                location = ', '.join(p for p in loc_parts if p) or 'Remote'
                workplace = item.get('workplace_type') or item.get('remote_policy') or ''
                posted = (item.get('source_posted_at') or item.get('first_seen_at') or '')[:10]
                jobs.append({
                    'company': _bluedoor_company_name(item),
                    'title': item.get('title', ''),
                    'url': item.get('apply_url') or item.get('source_url', ''),
                    'source': 'Bluedoor',
                    'location': location,
                    'tags': ', '.join(t for t in [item.get('department'), workplace] if t),
                    'posted_date': posted,
                    '_bd_job_id': item.get('job_id'),   # for late EMEA verification
                    '_bd_country': item.get('country') or '',
                })

            cursor = (payload.get('meta') or {}).get('next_cursor')
            if not cursor:
                break
        time.sleep(0.3)  # be polite

    print(f"  Bluedoor: {len(jobs)} EMEA remote jobs fetched")
    return jobs


# ── Bluedoor EMEA description verification ──
# Bluedoor's structured `country` tag is loose (e.g. a "global remote" role can be
# tagged with the company's HQ country). After role/geo filtering we fetch the real
# job description for the few survivors and (a) drop hard non-EMEA roles, (b) note
# when the true remote scope is broader than the country tag — surfaced in the digest.

# Phrases meaning the role is genuinely open beyond its tagged country
_BD_GLOBAL_SCOPE = [
    'work from anywhere', 'anywhere in the world', 'fully distributed',
    'globally remote', 'global remote', 'remote, global', 'work from any country',
    'work remotely from anywhere', 'anywhere in the globe',
]
_BD_EMEA_SCOPE = [
    'anywhere in europe', 'remote within europe', 'remote in europe',
    'europe-based', 'based anywhere in europe', 'eu remote', 'emea',
    'european time zone', 'cet timezone', 'cet time zone', 'within the eu',
]
# Hard residency clauses that make a role NOT workable from EMEA → drop
_BD_HARD_NON_EMEA = [
    'must be based in the united states', 'must reside in the united states',
    'us residents only', 'must be located in the united states',
    'must be a us citizen', 'us work authorization required',
    'us-based candidates only', 'within the united states only',
    'based in the united states only', 'must be based in the us',
    'must be located in canada', 'canada residents only',
    'must be based in india', 'latam only',
]


def _bluedoor_verify(job):
    """Fetch a kept Bluedoor job's description and decide keep/drop + annotate.

    Returns True to keep, False to drop. Drops only on an explicit hard non-EMEA
    residency clause. Sets job['location_note'] when the real remote scope is
    broader than the country tag. On any fetch failure, keeps the job (no note).
    """
    jid = job.get('_bd_job_id')
    if not jid:
        return True
    try:
        resp = requests.get(f'{BLUEDOOR_BASE}/jobs/{jid}', params={'include': 'description'},
                            headers={'User-Agent': 'RemoteJobSearch/1.0'}, timeout=15)
        resp.raise_for_status()
        desc = (resp.json().get('data', {}).get('description_text') or '').lower()
    except Exception as e:
        print(f"  Bluedoor verify error ({job.get('company','?')}): {e}")
        return True  # never silently drop on fetch failure
    if not desc:
        return True

    # (a) hard exclusion — not workable from EMEA
    for clause in _BD_HARD_NON_EMEA:
        if clause in desc:
            print(f"  Bluedoor DROP (non-EMEA): {job['company']} — '{clause}'")
            return False

    # (b) note when true scope is broader than a single-country tag
    tag = (job.get('_bd_country') or '').strip()
    scope = ''
    if any(p in desc for p in _BD_GLOBAL_SCOPE):
        scope = 'global remote'
    elif any(p in desc for p in _BD_EMEA_SCOPE):
        scope = 'Europe/EMEA remote'
    if scope and tag and tag.lower() not in ('', 'remote') and scope.split()[0] not in tag.lower():
        job['location_note'] = f"Tagged {tag} -> actually {scope}"
    return True


def verify_bluedoor_jobs(jobs):
    """Run EMEA description verification on kept Bluedoor jobs only (cheap: small set)."""
    bd = [j for j in jobs if j.get('source') == 'Bluedoor']
    if not bd:
        return jobs
    out, dropped, noted = [], 0, 0
    for job in jobs:
        if job.get('source') == 'Bluedoor':
            if not _bluedoor_verify(job):
                dropped += 1
                continue
            if job.get('location_note'):
                noted += 1
            time.sleep(0.2)  # be polite
        out.append(job)
    print(f"  Bluedoor verify: checked {len(bd)}, dropped {dropped}, annotated {noted}")
    return out


# Known EMEA tech companies → country (for enriching "Remote" listings)
KNOWN_COMPANY_COUNTRIES = {
    'tines': 'Ireland', 'intercom': 'Ireland', 'stripe': 'Ireland/US',
    'personio': 'Germany', 'celonis': 'Germany', 'contentful': 'Germany',
    'sap': 'Germany', 'delivery hero': 'Germany', 'zalando': 'Germany',
    'trustpilot': 'Denmark', 'unity': 'Denmark', 'pleo': 'Denmark',
    'klarna': 'Sweden', 'spotify': 'Sweden', 'king': 'Sweden',
    'adyen': 'Netherlands', 'booking.com': 'Netherlands', 'elastic': 'Netherlands',
    'revolut': 'UK', 'monzo': 'UK', 'wise': 'UK', 'deliveroo': 'UK',
    'datadog': 'France/US', 'criteo': 'France', 'doctolib': 'France',
    'alan': 'France', 'sorare': 'France', 'ledger': 'France',
    'veriff': 'Estonia', 'pipedrive': 'Estonia',
    'grafana labs': 'Sweden/Global', 'canonical': 'UK/Global',
    'gitlab': 'Global', 'automattic': 'Global',
    'meta': 'US', 'google': 'US', 'amazon': 'US', 'apple': 'US',
    'microsoft': 'US', 'netflix': 'US', 'uber': 'US', 'airbnb': 'US',
    'coinbase': 'US', 'robinhood': 'US', 'figma': 'US',
    'upstart': 'US', 'akamai': 'US', 'pnc': 'US', 'chainguard': 'US',
    'hashicorp': 'US', 'cloudflare': 'US', 'twilio': 'US',
}


def enrich_job_location(job):
    """Add company country info to jobs with vague 'Remote' location."""
    if job['location'].strip().lower() in ('remote', '', 'worldwide'):
        company_lower = job['company'].lower().strip()
        # Check known company lookup
        for known, country in KNOWN_COMPANY_COUNTRIES.items():
            if known in company_lower or company_lower in known:
                job['tags'] = f"{country}" + (f" | {job['tags']}" if job['tags'] else '')
                break
        # Check salary currency in tags/description for hints
        if '$' in job.get('tags', '') and '€' not in job.get('tags', ''):
            if not any(c in job.get('tags', '').lower() for c in ['ireland', 'uk', 'germany', 'france', 'global']):
                job['tags'] = (job['tags'] + ' | Likely US' if job['tags'] else 'Likely US')
    return job


# ============= FILTER, SORT & DEDUP =============

# US indicators in location text
US_INDICATORS = [
    'united states', 'usa', 'us ', 'u.s.', 'america',
    'new york', 'san francisco', 'los angeles', 'seattle', 'chicago',
    'austin', 'boston', 'denver', 'miami', 'washington',
    'california', 'texas', ', ny', ', ca', ', wa',
    'toronto', 'vancouver', 'canada',
]

# EMEA indicators that allow a job to pass even if US is also mentioned
EMEA_SIGNALS = [
    'anywhere', 'worldwide', 'global', 'emea', 'europe', 'eu',
    'france', 'paris', 'uk', 'germany', 'netherlands',
    'belgium', 'spain', 'italy', 'switzerland', 'austria',
    'sweden', 'norway', 'denmark', 'portugal', 'ireland',
    'poland', 'czech', 'united kingdom', 'london',
    'romania', 'hungary', 'greece', 'finland', 'croatia',
    'luxembourg', 'berlin', 'amsterdam', 'barcelona', 'munich',
    'dublin', 'lisbon', 'warsaw', 'prague', 'vienna', 'brussels',
    'gmt', 'cet', 'cest', 'utc+1', 'utc+2',
]


def filter_jobs(jobs):
    """Filter jobs by role keywords, exclude irrelevant roles, and check location."""
    filtered = []
    for job in jobs:
        title_lower = job['title'].lower()
        location_lower = job['location'].lower()
        tags_lower = job['tags'].lower()
        search_text = f"{title_lower} {tags_lower}"

        # Check REJECTED_REMOTE_LIST
        company_lower = job['company'].lower()
        blocklisted = False
        for bl_company, bl_title in REJECTED_REMOTE_LIST:
            if bl_company and bl_company not in company_lower:
                continue
            if bl_title == '' or bl_title in title_lower:
                blocklisted = True
                break
        if blocklisted:
            continue

        # Exclude roles that need ML/data science, are frontend, devops, etc.
        if any(ex in search_text for ex in ROLE_EXCLUDE):
            continue

        # Check primary role keywords
        matches_primary = any(kw in search_text for kw in ROLE_KEYWORDS)

        # Check secondary Python keyword
        matches_python = any(kw in search_text for kw in PYTHON_SECONDARY_KEYWORDS)

        if matches_python and not matches_primary:
            # Python-only role: only allow if it ALSO mentions Java/backend
            has_java_backend = any(sig in search_text for sig in JAVA_BACKEND_SIGNALS)
            if not has_java_backend:
                continue
        elif not matches_primary:
            # No keyword match at all
            continue

        # Strictly exclude jobs only targeting excluded regions
        # (skip for EMEA-searched sources — they were fetched with EMEA keywords)
        loc_tags = f"{location_lower} {tags_lower}"
        if job.get('source') not in EMEA_SEARCHED_SOURCES:
            if any(ex in loc_tags for ex in LOCATION_EXCLUDE):
                continue

        # Exclude US flag emoji
        if '\U0001f1fa\U0001f1f8' in job['location']:
            continue

        # Check for US/Canada mentions in location OR title
        loc_title = f"{location_lower} {title_lower}"
        has_us = any(us in loc_title for us in US_INDICATORS)

        if has_us:
            # Only keep if ALSO mentions an EMEA-compatible location
            has_emea = any(emea in loc_tags for emea in EMEA_SIGNALS)
            if not has_emea:
                continue

        # EU-focused sources (e.g., Arbeitnow) pass without explicit EMEA location
        if job.get('source') in EU_FOCUSED_SOURCES:
            filtered.append(job)
            continue

        # Must have an EMEA-compatible location signal
        has_emea_loc = any(inc in loc_tags for inc in LOCATION_INCLUDE)
        if has_emea_loc:
            filtered.append(job)
            continue

        # Relaxed sources: allow "Remote" with no region if no US indicators
        # Also exclude if enrichment tagged it as US company
        enriched_us = 'likely us' in tags_lower or tags_lower.startswith('us ')
        if job.get('source') in RELAXED_LOCATION_SOURCES and not has_us and not enriched_us:
            filtered.append(job)
            continue

    return filtered


def get_location_tier(job):
    """Return location priority tier (0=Paris, 1=France, 2=EMEA/CET, 3=UK, 4=global)."""
    loc = job['location'].lower()
    for keywords, tier in LOCATION_PRIORITY:
        if any(kw in loc for kw in keywords):
            return tier
    return 5  # Unknown location goes last


def is_explicitly_remote(job):
    """Check if a job explicitly mentions remote in title or location."""
    text = f"{job['title'].lower()} {job['location'].lower()}"
    return 'remote' in text or 'full remote' in text or 'télétravail' in text


def sort_jobs(jobs):
    """Sort by location tier, then remote-first within tier, then by date descending."""
    return sorted(jobs, key=lambda j: (
        get_location_tier(j),
        0 if is_explicitly_remote(j) else 1,  # remote first within tier
        j['posted_date'][::-1],
    ))


def dedup_jobs(jobs):
    """Remove duplicates by (company, title) pair."""
    seen = set()
    unique = []
    for job in jobs:
        key = (job['company'].lower().strip(), job['title'].lower().strip())
        if key not in seen:
            seen.add(key)
            unique.append(job)
    return unique


# ============= JOB HISTORY TRACKING =============

HISTORY_FILE = os.path.join(os.path.dirname(__file__), 'previous_jobs.json')


def load_previous_jobs():
    """Load previous job keys from history file."""
    try:
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return set(tuple(k) for k in json.load(f))
    except (FileNotFoundError, json.JSONDecodeError):
        return set()


def save_current_jobs(jobs):
    """Save current job keys to history file for next run comparison."""
    keys = [(j['company'].lower().strip(), j['title'].lower().strip()) for j in jobs]
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(keys, f)


def mark_new_jobs(jobs, previous_keys):
    """Mark each job as new or existing based on previous run."""
    for job in jobs:
        key = (job['company'].lower().strip(), job['title'].lower().strip())
        job['is_new'] = key not in previous_keys
    return jobs


# ============= EXCEL DUMP =============

EXCEL_HEADERS = ['Company', 'Role', 'URL', 'Source', 'Location', 'Tags', 'Posted', 'New?']
REMOTE_EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'remote.xlsx')


def _safe_str(val):
    """Return a string safe for Excel cells — replaces chars that Windows codepage can't encode."""
    if not isinstance(val, str):
        return val
    return (val
            .replace('\u2192', '->')
            .replace('\u2190', '<-')
            .replace('\u2022', '*')
            .replace('\u2013', '-')
            .replace('\u2014', '-')
            .encode('cp1252', errors='replace').decode('cp1252'))


def dump_to_excel(jobs):
    """Append new remote jobs to remote.xlsx — never removes existing rows.

    Jobs already present (matched by company+title) are skipped.
    Creates the file with a header row if it doesn't exist yet.
    Handles PermissionError (file open in Excel) via temp file.
    """
    temp_out = None
    try:
        # Load existing file or create fresh
        if os.path.exists(REMOTE_EXCEL_FILE):
            try:
                wb = openpyxl.load_workbook(REMOTE_EXCEL_FILE)
            except PermissionError:
                temp_fd, temp_in = tempfile.mkstemp(suffix='.xlsx')
                os.close(temp_fd)
                shutil.copy2(REMOTE_EXCEL_FILE, temp_in)
                wb = openpyxl.load_workbook(temp_in)
                os.remove(temp_in)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'remote'
            ws.append(EXCEL_HEADERS)
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)

        # Build set of existing (company, title) keys
        existing_keys = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            company, title = str(row[0] or '').lower().strip(), str(row[1] or '').lower().strip()
            existing_keys.add((company, title))

        # Append only new jobs
        added = 0
        for job in jobs:
            key = (job['company'].lower().strip(), job['title'].lower().strip())
            if key in existing_keys:
                continue
            new_flag = 'NEW' if job.get('is_new') else ''
            location_cell = job['location']
            if job.get('location_note'):
                location_cell = f"{location_cell} ({job['location_note']})"
            ws.append([
                _safe_str(job['company']),
                _safe_str(job['title']),
                job['url'],
                _safe_str(job['source']),
                _safe_str(location_cell),
                _safe_str(job['tags']),
                _safe_str(job['posted_date']),
                new_flag,
            ])
            url_cell = ws.cell(row=ws.max_row, column=3)
            url_cell.hyperlink = job['url']
            url_cell.style = 'Hyperlink'
            existing_keys.add(key)
            added += 1

        # Auto-width
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

        # Save
        try:
            wb.save(REMOTE_EXCEL_FILE)
            print(f"Excel dump: {added} new jobs appended -> remote.xlsx (total rows: {ws.max_row - 1})")
        except PermissionError:
            temp_fd, temp_out = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)
            wb.save(temp_out)
            shutil.copy2(temp_out, REMOTE_EXCEL_FILE)
            print(f"Excel dump (via temp): {added} new jobs appended -> remote.xlsx")

    except Exception as e:
        print(f"Warning: Excel dump failed - {e}")
    finally:
        if temp_out and os.path.exists(temp_out):
            os.remove(temp_out)


# ============= HTML EMAIL =============

TIER_LABELS = {0: 'Paris', 1: 'France', 2: 'EMEA / CET', 3: 'UK', 4: 'Global / Remote', 5: 'Remote (Region Unspecified)'}

def build_html(jobs, new_count=0, total_unchanged=False):
    """Build styled HTML email with job listings grouped by location tier."""
    now = datetime.now().strftime('%Y-%m-%d %H:%M')
    date_str = datetime.now().strftime('%Y-%m-%d')

    # Banner for no-change or new jobs count
    banner_html = ''
    if total_unchanged:
        banner_html = '<div style="background:#fff3cd;color:#856404;padding:8px 12px;border-radius:4px;margin-bottom:8px;font-size:12px;font-weight:bold;text-align:center;">No new jobs since last run — all listings unchanged.</div>'
    elif new_count > 0:
        banner_html = f'<div style="background:#d4edda;color:#155724;padding:8px 12px;border-radius:4px;margin-bottom:8px;font-size:12px;font-weight:bold;text-align:center;">🆕 {new_count} new job(s) since last run — highlighted in green below.</div>'

    rows_html = ''
    current_tier = None
    for job in jobs:
        tier = get_location_tier(job)
        if tier != current_tier:
            current_tier = tier
            label = TIER_LABELS.get(tier, 'Other')
            rows_html += f'                <tr style="background:#e8f5e9;font-weight:bold;"><td colspan="6" style="padding:4px 6px;font-size:11px;color:#2c3e50;">📍 {label}</td></tr>\n'

        is_new = job.get('is_new', False)
        row_style = ' style="background:#d4edda;"' if is_new else ''
        new_badge = ' <span style="background:#28a745;color:white;padding:1px 4px;border-radius:3px;font-size:9px;font-weight:bold;">NEW</span>' if is_new else ''

        note_html = ''
        if job.get('location_note'):
            note_html = f'<br><span style="font-size:9px;color:#e67e22;font-weight:bold;">✅ {job["location_note"]}</span>'

        rows_html += f"""                <tr{row_style}>
                    <td><strong>{job['company']}</strong>{new_badge}</td>
                    <td><a href="{job['url']}" style="color: #3498db; text-decoration: underline;">{job['title']}</a></td>
                    <td>{job['source']}</td>
                    <td>{job['location']}{note_html}<br><span style="font-size:9px;color:#7f8c8d;">{job['tags']}</span></td>
                    <td>{fit_badge_html(job.get('fit'))}</td>
                    <td>{job['posted_date']}</td>
                </tr>
"""

    if not jobs:
        rows_html = '<tr><td colspan="6" style="text-align:center;padding:20px;color:#7f8c8d;">No matching remote roles found this run.</td></tr>\n'

    sources = 'RemoteOK, Remotive, WWR, Jobicy, LinkedIn FR, LinkedIn Global, Bluedoor'
    html = f"""
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 11px; line-height: 1.2; color: #2c3e50; max-width: 1600px; margin: 0; padding: 8px; background: #f8f9fa; }}
            h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 5px; margin: 5px 0; font-size: 18px; }}
            .header-info {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 6px 10px; border-radius: 4px; margin-bottom: 8px; font-size: 11px; }}
            .header-info p {{ margin: 2px 0; }}
            table {{ border-collapse: collapse; width: 100%; background: white; box-shadow: 0 1px 3px rgba(0,0,0,0.08); font-size: 11px; margin: 5px 0; }}
            th {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; font-weight: bold; padding: 5px 6px; text-align: left; font-size: 10px; }}
            td {{ padding: 4px 6px; border-bottom: 1px solid #ecf0f1; vertical-align: top; }}
            tr:hover {{ background-color: #f8f9fa; }}
            .footer {{ text-align: center; color: #7f8c8d; font-size: 10px; margin-top: 12px; padding-top: 8px; border-top: 2px solid #ecf0f1; }}
        </style>
    </head>
    <body>
        <h1>🌍 Remote Roles — {date_str}</h1>

        <div class="header-info">
            <p>📅 {now} | 📊 {len(jobs)} matching roles | 🔍 Sorted: Paris → France → EMEA/CET → UK → Global | Sources: {sources}</p>
        </div>

        {banner_html}

        <table>
            <thead>
                <tr>
                    <th width="16%">Company</th>
                    <th width="26%">Role</th>
                    <th width="9%">Source</th>
                    <th width="23%">Location / Tags</th>
                    <th width="11%">Fit</th>
                    <th width="15%">Posted</th>
                </tr>
            </thead>
            <tbody>
{rows_html}            </tbody>
        </table>

        <div class="footer">
            <p>📊 {len(jobs)} remote roles | Sources: {sources} | 🔄 Next: in 2 days at 12:00 CET</p>
        </div>
    </body>
    </html>
    """
    return html


def send_email(html_content):
    """Send the HTML email using config."""
    try:
        message = MIMEMultipart("alternative")
        message["Subject"] = f"Remote Roles — {datetime.now().strftime('%Y-%m-%d')}"
        message["From"] = EMAIL_CONFIG['sender_email']
        message["To"] = EMAIL_CONFIG['recipient_email']

        message.attach(MIMEText(html_content, "html"))

        with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port']) as server:
            server.starttls()
            server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
            server.send_message(message)

        print("SUCCESS: Remote jobs email sent")
        return True
    except Exception as e:
        print(f"ERROR sending email: {e}")
        return False


# ============= MAIN =============

def main(no_save=False):
    print("=== Remote Job Search ===")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    if no_save:
        print("Mode: TEST (previous_jobs.json will NOT be updated)")

    # Fetch from all sources
    print("Fetching jobs...")
    all_jobs = []
    all_jobs.extend(fetch_remoteok())
    all_jobs.extend(fetch_remotive())
    # Arbeitnow removed — low quality results
    # all_jobs.extend(fetch_arbeitnow())
    all_jobs.extend(fetch_weworkremotely())
    all_jobs.extend(fetch_jobicy())
    all_jobs.extend(fetch_linkedin_france())
    all_jobs.extend(fetch_linkedin_global())
    if BLUEDOOR_ENABLED:
        all_jobs.extend(fetch_bluedoor())
    else:
        print("  Bluedoor: disabled (BLUEDOOR_ENABLED = False)")
    print(f"Total fetched: {len(all_jobs)}")

    # Enrich location info for vague "Remote" listings
    all_jobs = [enrich_job_location(job) for job in all_jobs]

    # Filter and dedup
    filtered = filter_jobs(all_jobs)
    print(f"After role/location filter: {len(filtered)}")

    deduped = dedup_jobs(filtered)
    print(f"After dedup: {len(deduped)}")

    # Verify Bluedoor survivors against their real descriptions (EMEA + scope note)
    deduped = verify_bluedoor_jobs(deduped)
    print(f"After Bluedoor EMEA verify: {len(deduped)}")

    # Sort: Paris → France → EMEA/CET → UK → Global, remote-first within tier
    sorted_jobs = sort_jobs(deduped)

    # Fit scoring — one batched Gemini call for all remote jobs
    if _FIT_AVAILABLE and sorted_jobs:
        print(f"Scoring job fit for {len(sorted_jobs)} remote jobs...")
        items = [{'title': j['title'], 'company': j['company']} for j in sorted_jobs]
        fit_scores = score_fit_batch(items)
        for job, fit in zip(sorted_jobs, fit_scores):
            job['fit'] = fit
        scored = sum(1 for f in fit_scores if f)
        print(f"  Fit scores added: {scored}/{len(sorted_jobs)}")

    # Compare with previous run to detect new jobs
    previous_keys = load_previous_jobs()
    sorted_jobs = mark_new_jobs(sorted_jobs, previous_keys)
    new_count = sum(1 for j in sorted_jobs if j.get('is_new'))
    total_unchanged = len(previous_keys) > 0 and new_count == 0

    if total_unchanged:
        print(f"No new jobs since last run (all {len(sorted_jobs)} unchanged)")
    else:
        print(f"New jobs: {new_count} / {len(sorted_jobs)} total")

    # Save current jobs for next run comparison (skip on test runs)
    if no_save:
        print("Skipping previous_jobs.json update (--no-save)")
    else:
        save_current_jobs(sorted_jobs)

    # Dump to Excel tracker ('remote' sheet)
    dump_to_excel(sorted_jobs)

    # Build HTML and send
    html = build_html(sorted_jobs, new_count=new_count, total_unchanged=total_unchanged)
    send_email(html)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--no-save', action='store_true',
                        help='Skip updating previous_jobs.json (use for test runs to keep NEW flags intact for the next scheduled run)')
    args = parser.parse_args()
    main(no_save=args.no_save)
