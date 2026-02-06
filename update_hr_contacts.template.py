"""
Update List.xlsx with potentialHR contact column
- Searches for recruiter/talent acquisition contacts per company
- Names only (no expertise/roles) with clickable LinkedIn hyperlinks
- Column placed before "Other comments"

Usage:
1. Copy this file to update_hr_contacts.py
2. Fill in HR_CONTACTS with your researched contacts
3. Run: python update_hr_contacts.py
"""
import openpyxl
from openpyxl.styles import Font, Alignment
from copy import copy

# Import tracker path from config
try:
    from config import TRACKER_FILE
except ImportError:
    print("ERROR: config.py not found!")
    exit(1)

# HR contacts: company -> list of (name, linkedin_url) tuples
# Fill this in with your own research (e.g. search LinkedIn for
# "CompanyName talent acquisition recruiter YourCity site:linkedin.com/in")
HR_CONTACTS = {
    # Example entries:
    "Example Corp": [
        ("Jane Smith", "https://www.linkedin.com/in/janesmith/"),
        ("John Doe", "https://www.linkedin.com/in/johndoe/"),
    ],
    "Another Company": [
        ("Recruiter Name", "https://www.linkedin.com/in/recruitername/"),
    ],
    # Add your companies here...
}


def find_company_contacts(company):
    """Look up HR contacts with case-insensitive/partial matching"""
    contacts = HR_CONTACTS.get(company)
    if contacts:
        return contacts
    for key, value in HR_CONTACTS.items():
        if key.lower() == company.lower() or key.lower() in company.lower() or company.lower() in key.lower():
            return value
    return None


def update_excel():
    """Rewrite potentialHR column: names + clickable links, before Other comments"""
    wb = openpyxl.load_workbook(TRACKER_FILE)
    ws = wb.active

    header_row = [cell.value for cell in ws[1]]
    print(f"Current headers: {header_row}")

    # Step 1: Remove existing potentialHR column if it exists
    hr_existing_col = None
    for idx, header in enumerate(header_row):
        if header and 'potentialhr' in str(header).lower().replace(' ', ''):
            hr_existing_col = idx + 1
            break

    if hr_existing_col:
        print(f"Removing old potentialHR column at position {hr_existing_col}")
        ws.delete_cols(hr_existing_col)
        header_row = [cell.value for cell in ws[1]]
        print(f"Headers after removal: {header_row}")

    # Step 2: Find "Other comments" column and insert before it
    other_comments_col = None
    for idx, header in enumerate(header_row):
        if header and 'other comments' in str(header).lower():
            other_comments_col = idx + 1
            break

    if other_comments_col:
        hr_col = other_comments_col
        print(f"Inserting potentialHR contact at column {hr_col} (before 'Other comments')")
        ws.insert_cols(hr_col)
    else:
        hr_col = len(header_row) + 1
        print(f"'Other comments' not found, adding at column {hr_col}")

    # Step 3: Set header
    header_cell = ws.cell(row=1, column=hr_col, value="potentialHR contact")
    try:
        source_cell = ws.cell(row=1, column=hr_col + 1)
        if source_cell.font:
            header_cell.font = copy(source_cell.font)
        if source_cell.fill:
            header_cell.fill = copy(source_cell.fill)
        if source_cell.alignment:
            header_cell.alignment = copy(source_cell.alignment)
    except Exception:
        pass

    # Step 4: Populate with clickable hyperlinks (name only)
    link_font = Font(color="0563C1", underline="single", size=10)
    updated = 0
    skipped = 0

    for row_idx in range(2, ws.max_row + 1):
        company_cell = ws.cell(row=row_idx, column=1)
        company = str(company_cell.value).strip() if company_cell.value else ''

        if not company or 'Program/Product' in company:
            continue

        contacts = find_company_contacts(company)

        if contacts:
            if len(contacts) == 1:
                name, url = contacts[0]
                cell = ws.cell(row=row_idx, column=hr_col)
                cell.value = name
                cell.hyperlink = url
                cell.font = link_font
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                # Multiple contacts: concatenate HYPERLINK formulas
                parts = []
                for name, url in contacts:
                    parts.append(f'HYPERLINK("{url}","{name}")')
                formula = " & CHAR(10) & ".join(parts)
                cell = ws.cell(row=row_idx, column=hr_col)
                cell.value = f"={formula}"
                cell.font = link_font
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            updated += 1
        else:
            skipped += 1
            print(f"  Row {row_idx}: {company} - no HR contact found")

    ws.column_dimensions[openpyxl.utils.get_column_letter(hr_col)].width = 40

    wb.save(TRACKER_FILE)
    wb.close()
    print(f"\nDone! Updated {updated} companies, skipped {skipped}")
    print(f"File saved: {TRACKER_FILE}")


if __name__ == "__main__":
    update_excel()
